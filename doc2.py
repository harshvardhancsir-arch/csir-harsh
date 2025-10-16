# file: app.py
"""
Streamlit app — Local Research Paper Extractor
(Updated to fix IllegalCharacterError and add progress visualization)
"""

from pathlib import Path
import re
import io
import tempfile
import time
import traceback

import pandas as pd
import streamlit as st
from pdfminer.high_level import extract_text
from pdf2image import convert_from_path
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Config ----------
TITLE_MIN_LEN = 15
MAX_TITLE_LINES = 3
HEADINGS = [
    "abstract", "introduction", "background", "literature review", "related work", "literature", "review",
    "methodology", "methods", "materials and methods", "experimental", "method", "approach", "procedure",
    "results", "findings", "discussion", "conclusion", "conclusions", "summary", "implications",
    "references", "bibliography", "works cited"
]

# Roman numerals for section detection
ROMAN_NUMERALS = [
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV"
]

# ---------- Utility functions ----------
def clean_text_for_excel(text: str) -> str:
    """Remove illegal Unicode characters for Excel."""
    if not isinstance(text, str):
        return text
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = Path(tmp.name)
    text = ""
    try:
        text = extract_text(str(tmp_path))
    except Exception:
        text = ""
    if len(text.strip()) < 200:
        try:
            pages = convert_from_path(str(tmp_path), dpi=200)
            text = "\n\n".join(pytesseract.image_to_string(p.convert("L")) for p in pages)
        except Exception as e:
            text = f"[OCR_FAILED] {e}"
    tmp_path.unlink(missing_ok=True)
    return text

def normalize_text(text: str) -> str:
    text = text.replace("\r", "\n")
    text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
    return text.strip()

def find_title_and_authors(text: str) -> dict:
    """
    Improved title detection that looks for:
    1. Text in ALL CAPS (common for titles)
    2. Centered text patterns
    3. Text with proper title formatting
    4. Text that's clearly separated from other content
    """
    
    # Split text into lines and analyze each line
    lines = text.split('\n')
    title_candidates = []
    
    # Look for title patterns in the first 20 lines
    for i, line in enumerate(lines[:20]):
        line = line.strip()
        if not line or len(line) < 10:  # Skip empty or very short lines
            continue
            
        # Skip lines that are clearly not titles
        if any(skip_word in line.lower() for skip_word in [
            'abstract', 'introduction', 'keywords', 'author', 'email', 
            'university', 'college', 'department', 'institute', 'correspondence',
            'received', 'accepted', 'published', 'doi:', 'http', 'www',
            'business school', 'school', 'faculty', 'professor', 'dr.', 'prof.',
            'phd', 'mba', 'msc', 'bsc', 'ma', 'ba', 'corresponding author',
            'affiliation', 'address', 'phone', 'fax', 'email:', 'e-mail',
            'china', 'usa', 'uk', 'germany', 'france', 'japan', 'india',
            'beijing', 'shanghai', 'london', 'new york', 'tokyo', 'berlin'
        ]):
            continue
        
        # Calculate title score based on multiple factors
        score = 0
        
        # Factor 1: Length (titles are usually 20-200 characters)
        if 20 <= len(line) <= 200:
            score += 2
        elif 10 <= len(line) <= 300:
            score += 1
            
        # Factor 2: ALL CAPS (very common for academic titles)
        if line.isupper() and len(line) > 15:
            score += 3
            
        # Factor 3: Title case (First Letter Of Each Word Capitalized)
        if line.istitle() and len(line) > 20:
            score += 2
            
        # Factor 3.5: Mixed case (some words capitalized, some not - common in titles)
        words = line.split()
        if len(words) > 2:
            capitalized_words = sum(1 for word in words if word[0].isupper() and len(word) > 1)
            if capitalized_words >= len(words) * 0.5:  # At least 50% of words capitalized
                score += 1
            
        # Factor 4: Contains academic keywords
        academic_words = [
            'study', 'analysis', 'approach', 'method', 'model', 'system',
            'framework', 'algorithm', 'technique', 'evaluation', 'assessment',
            'investigation', 'research', 'development', 'implementation',
            'application', 'comparison', 'review', 'survey', 'survey'
        ]
        if any(word in line.lower() for word in academic_words):
            score += 1
            
        # Factor 5: Position (titles are usually early in the document)
        if i < 5:
            score += 2
        elif i < 10:
            score += 1
            
        # Factor 6: Not too many numbers (titles rarely have many numbers)
        number_count = sum(1 for c in line if c.isdigit())
        if number_count < len(line) * 0.1:  # Less than 10% numbers
            score += 1
            
        # Factor 7: Not too many special characters
        special_count = sum(1 for c in line if not c.isalnum() and not c.isspace())
        if special_count < len(line) * 0.2:  # Less than 20% special chars
            score += 1
            
        # Factor 8: Line is not too long (titles are usually concise)
        if len(line) < 100:
            score += 1
            
        # Factor 9: Contains common title words
        title_words = [
            'a', 'an', 'the', 'of', 'in', 'on', 'at', 'to', 'for', 'with',
            'by', 'from', 'and', 'or', 'but', 'using', 'based', 'towards'
        ]
        if any(word in line.lower() for word in title_words):
            score += 0.5
            
        # Factor 10: Line is followed by author information
        if i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            if any(author_word in next_line.lower() for author_word in [
                'author', 'by', 'university', 'college', 'department'
            ]):
                score += 2
        
        # Penalty factors - reduce score for lines that look like author info
        penalty = 0
        
        # Penalty 1: Contains numbers at the beginning (like "1 Business School")
        if re.match(r'^\d+\s+', line):
            penalty += 3
            
        # Penalty 2: Very short lines (likely author names)
        if len(line) < 30:
            penalty += 1
            
        # Penalty 3: Contains location names (cities, countries)
        location_words = [
            'china', 'usa', 'uk', 'germany', 'france', 'japan', 'india',
            'beijing', 'shanghai', 'london', 'new york', 'tokyo', 'berlin',
            'sichuan', 'california', 'texas', 'florida', 'ontario'
        ]
        if any(loc in line.lower() for loc in location_words):
            penalty += 2
            
        # Penalty 4: Contains academic titles/degrees
        degree_words = ['professor', 'dr.', 'prof.', 'phd', 'mba', 'msc', 'bsc']
        if any(degree in line.lower() for degree in degree_words):
            penalty += 2
            
        # Penalty 5: Contains institutional words
        institutional_words = ['school', 'university', 'college', 'institute', 'faculty']
        if any(inst in line.lower() for inst in institutional_words):
            penalty += 2
        
        # Apply penalty
        score -= penalty
        
        # Only consider lines with decent scores (increased threshold)
        if score >= 4:  # Increased from 3 to 4
            title_candidates.append((line, score, i))
    
    # Sort by score (highest first) and position (earlier first)
    title_candidates.sort(key=lambda x: (-x[1], x[2]))
    
    # Get the best title candidate
    if title_candidates:
        best_title = title_candidates[0][0]
        
        # Final validation - make sure it's not author information
        if (re.match(r'^\d+\s+', best_title) or  # Starts with number
            any(author_word in best_title.lower() for author_word in [
                'business school', 'school', 'university', 'college', 'institute',
                'professor', 'dr.', 'prof.', 'phd', 'mba', 'msc', 'bsc',
                'china', 'usa', 'uk', 'germany', 'france', 'japan', 'india',
                'sichuan', 'beijing', 'shanghai', 'london', 'new york'
            ]) or
            len(best_title) < 20 or  # Too short
            best_title.isupper() and len(best_title) < 30):  # Short all-caps (likely author name)
            # If best candidate is author info, try next candidate
            if len(title_candidates) > 1:
                best_title = title_candidates[1][0]
                title_idx = title_candidates[1][2]
            else:
                title = None
                return {"title": title, "authors": None}
        else:
            title_idx = title_candidates[0][2]
        
        if best_title:
            # Try to extend title to multiple lines if needed
            title_lines = [best_title]
            
            # Check if next lines might be part of the title
            for j in range(1, 3):  # Check next 2 lines
                if title_idx + j < len(lines):
                    next_line = lines[title_idx + j].strip()
                    if (len(next_line) > 10 and 
                        not any(skip_word in next_line.lower() for skip_word in [
                            'abstract', 'introduction', 'author', 'email', 'university',
                            'business school', 'school', 'professor', 'dr.', 'prof.'
                        ]) and
                        len(' '.join(title_lines + [next_line])) < 300):  # Total title not too long
                        title_lines.append(next_line)
                    else:
                        break
            
            title = ' '.join(title_lines)
        else:
            title = None
    else:
        title = None
    
    # Extract authors (look for lines after title)
    authors = None
    if title:
        title_end_idx = title_idx + len(title_lines)
        author_candidates = []
        
        # Look for author information in next 10 lines after title
        for i in range(title_end_idx, min(title_end_idx + 10, len(lines))):
            line = lines[i].strip()
            if not line:
                continue
                
            # Skip if it looks like abstract or other sections
            if any(skip_word in line.lower() for skip_word in [
                'abstract', 'introduction', 'keywords', 'doi:', 'http'
            ]):
                    break
                
            # Look for author patterns
            if (len(line) > 5 and 
                any(author_word in line.lower() for author_word in [
                    'university', 'college', 'department', 'institute', 'correspondence'
                ]) or
                ('@' in line and '.' in line) or  # Email pattern
                (len(line.split()) <= 5 and not line.isupper())):  # Short lines that aren't all caps
                author_candidates.append(line)
        
        if author_candidates:
            authors = ', '.join(author_candidates[:3])  # Take first 3 author lines
    
    return {"title": title, "authors": authors}

def find_roman_numeral_sections(text: str) -> dict:
    """Find sections with Roman numeral headings like 'VI. Conclusion'."""
    sections = {}
    
    # Pattern for Roman numeral headings - more comprehensive
    roman_pattern = r'^\s*([IVX]+)\.?\s+(conclusion|conclusions|discussion|summary|implications|findings|literature\s+review|related\s+work|methodology|methods|introduction|abstract|references|bibliography|literature|review)\s*[:\-\s]*\s*$'
    roman_re = re.compile(roman_pattern, flags=re.IGNORECASE | re.MULTILINE)
    matches = list(roman_re.finditer(text))
    
    print(f"Roman numeral matches found: {len(matches)}")
    for match in matches:
        print(f"Roman match: '{match.group(0).strip()}' -> section: '{match.group(2).strip().lower()}'")
    
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_name = m.group(2).strip().lower()
        content = text[start:end].strip()
        if content and len(content) > 50:
            sections[section_name] = content
            print(f"Added Roman section '{section_name}' with {len(content)} characters")
    
    return sections

def split_sections(text: str) -> dict:
    # More flexible pattern that handles various heading formats
    pattern = r"^\s*(?:{})\s*[:\-\s]*\s*$".format("|".join(re.escape(h) for h in HEADINGS))
    heading_re = re.compile(pattern, flags=re.IGNORECASE | re.MULTILINE)
    matches = list(heading_re.finditer(text))
    
    # Also look for numbered headings like "1. Introduction", "2. Methodology", etc.
    numbered_pattern = r"^\s*\d+\.?\s*(?:{})\s*[:\-\s]*\s*$".format("|".join(re.escape(h) for h in HEADINGS))
    numbered_re = re.compile(numbered_pattern, flags=re.IGNORECASE | re.MULTILINE)
    numbered_matches = list(numbered_re.finditer(text))
    
    # Look for Roman numeral headings
    roman_sections = find_roman_numeral_sections(text)
    
    # Combine and sort all matches
    all_matches = matches + numbered_matches
    all_matches.sort(key=lambda x: x.start())
    
    sections = {}
    for i, m in enumerate(all_matches):
        start = m.end()
        end = all_matches[i + 1].start() if i + 1 < len(all_matches) else len(text)
        section_name = m.group(0).strip().lower()
        # Clean up section name (remove numbers, colons, etc.)
        section_name = re.sub(r'^\d+\.?\s*', '', section_name)
        section_name = re.sub(r'[:\-\s]+$', '', section_name)
        content = text[start:end].strip()
        if content and len(content) > 50:  # Only include sections with substantial content
            sections[section_name] = content
    
    # Add Roman numeral sections (they take precedence)
    for key, value in roman_sections.items():
        if value and len(value) > 50:
            sections[key] = value
    
    return sections

def pick_section_content(sections, keys):
    """Find the best matching section content based on provided keys."""
    best_match = ""
    best_score = 0
    
    for k in keys:
        for sec, content in sections.items():
            if not content.strip():
                continue
                
            # Calculate match score based on how well the key matches the section name
            sec_lower = sec.lower()
            key_lower = k.lower()
            
            # Exact match gets highest score
            if key_lower == sec_lower:
                return content.strip()
            
            # Partial match gets medium score
            if key_lower in sec_lower or sec_lower in key_lower:
                score = len(key_lower) / len(sec_lower) if len(sec_lower) > 0 else 0
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
            
            # Word-based matching for compound terms
            key_words = set(key_lower.split())
            sec_words = set(sec_lower.split())
            if key_words.intersection(sec_words):
                score = len(key_words.intersection(sec_words)) / len(key_words)
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
    
    return best_match

def is_literature_review_content(content: str) -> bool:
    """Validate if content appears to be from a literature review section."""
    if not content or len(content.strip()) < 200:
        return False
    
    content_lower = content.lower()
    
    # Look for strong literature review indicators
    strong_indicators = [
        'literature review', 'related work', 'previous work', 'existing research',
        'prior research', 'earlier studies', 'recent studies', 'studies have shown',
        'research has shown', 'according to', 'authors have', 'scholars have',
        'researchers have', 'findings suggest', 'evidence shows', 'investigation',
        'analysis of', 'review of', 'state of the art', 'background research'
    ]
    
    # Look for weak indicators that might be in other sections
    weak_indicators = [
        'method', 'approach', 'procedure', 'experiment', 'data', 'results',
        'conclusion', 'summary', 'introduction', 'abstract', 'methodology'
    ]
    
    # Count strong indicators
    strong_count = sum(1 for indicator in strong_indicators if indicator in content_lower)
    
    # Count weak indicators (these should be minimal in literature review)
    weak_count = sum(1 for indicator in weak_indicators if indicator in content_lower)
    
    # Must have at least 3 strong indicators and no more than 2 weak indicators
    return strong_count >= 3 and weak_count <= 2

def is_methodology_content(content: str) -> bool:
    """Validate if content appears to be from a methodology section."""
    if not content or len(content.strip()) < 100:
        return False
    
    content_lower = content.lower()
    
    # Look for strong methodology indicators
    strong_method_indicators = [
        'methodology', 'methods', 'materials and methods', 'experimental', 'approach', 'procedure',
        'data collection', 'data analysis', 'experiment', 'experimental design', 'research design',
        'sampling', 'participants', 'subjects', 'procedure', 'protocol', 'technique', 'algorithm',
        'implementation', 'setup', 'configuration', 'parameters', 'variables', 'measurement',
        'instruments', 'tools', 'equipment', 'software', 'hardware', 'platform'
    ]
    
    # Look for weak indicators that might be in other sections
    weak_method_indicators = [
        'conclusion', 'summary', 'results', 'findings', 'discussion', 'introduction',
        'abstract', 'literature', 'background', 'related work', 'references'
    ]
    
    # Count strong methodology indicators
    strong_count = sum(1 for indicator in strong_method_indicators if indicator in content_lower)
    
    # Count weak indicators (these should be minimal in methodology)
    weak_count = sum(1 for indicator in weak_method_indicators if indicator in content_lower)
    
    # Must have at least 2 strong indicators and no more than 1 weak indicator
    return strong_count >= 2 and weak_count <= 1

def is_conclusion_content(content: str) -> bool:
    """Validate if content appears to be from a conclusion section."""
    if not content or len(content.strip()) < 100:
        return False
    
    content_lower = content.lower()
    
    # Look for strong conclusion indicators
    strong_concl_indicators = [
        'conclusion', 'conclude', 'summary', 'overall', 'finally', 'in summary',
        'results show', 'findings indicate', 'implications', 'recommendations',
        'future work', 'limitations', 'contribution', 'significance', 'impact',
        'this study', 'our research', 'we found', 'we conclude', 'we suggest',
        'in conclusion', 'to conclude', 'to summarize', 'in summary', 'overall',
        'the results', 'our findings', 'the study', 'this research', 'we have shown'
    ]
    
    # Look for weak indicators that might be in other sections
    weak_concl_indicators = [
        'method', 'approach', 'procedure', 'experiment', 'data', 'literature',
        'introduction', 'abstract', 'methodology', 'background', 'related work'
    ]
    
    # Count strong conclusion indicators
    strong_count = sum(1 for indicator in strong_concl_indicators if indicator in content_lower)
    
    # Count weak indicators (these should be minimal in conclusion)
    weak_count = sum(1 for indicator in weak_concl_indicators if indicator in content_lower)
    
    # Must have at least 2 strong indicators and no more than 1 weak indicator
    return strong_count >= 2 and weak_count <= 1

def extract_sections_by_content_patterns(text: str) -> dict:
    """Extract sections based on content patterns when headings are not clear."""
    sections = {}
    
    # Abstract pattern - usually starts with "Abstract" or similar and ends before Introduction
    abstract_match = re.search(r'(?i)(?:abstract|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:introduction|1\.|background|overview))', text, re.DOTALL)
    if abstract_match:
        sections['abstract'] = abstract_match.group(1).strip()
    
    # Introduction pattern - often starts with "Introduction" or "1."
    intro_match = re.search(r'(?i)(?:introduction|1\.)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|2\.|literature|related work))', text, re.DOTALL)
    if intro_match:
        sections['introduction'] = intro_match.group(1).strip()
    
    # Literature review pattern - more precise patterns including Roman numerals
    lit_patterns = [
        r'(?i)(?:[IVX]+\.?\s+)?(?:literature review|related work|previous work|background|state of the art)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach|procedure))',
        r'(?i)(?:2\.?\s*)?(?:literature review|related work|previous work)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach))',
        r'(?i)(?:literature|review|related work|previous work)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach|procedure))',
        # Specific Roman numeral patterns for literature review
        r'(?i)(?:II|III|IV|V)\.?\s+(?:literature review|related work|literature|review)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach|procedure))'
    ]
    
    for pattern in lit_patterns:
        lit_match = re.search(pattern, text, re.DOTALL)
        if lit_match:
            content = lit_match.group(1).strip()
            if len(content) > 200 and is_literature_review_content(content):
                sections['literature review'] = content
                break
    
    # Methodology pattern - often contains words like "method", "approach", "procedure"
    method_match = re.search(r'(?i)(?:methodology|methods|materials and methods|experimental|approach|procedure)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:results|findings|3\.|discussion))', text, re.DOTALL)
    if method_match:
        sections['methodology'] = method_match.group(1).strip()
    
    # Conclusion pattern - more precise patterns including Roman numerals
    concl_patterns = [
        r'(?i)(?:[IVX]+\.?\s+)?(?:conclusion|conclusions|discussion|summary|implications|findings)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
        r'(?i)(?:\d+\.?\s*)?(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
        r'(?i)(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|$))',
        # Specific Roman numeral patterns
        r'(?i)(?:VI|VII|VIII|IX|X|XI|XII)\.?\s+(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))'
    ]
    
    for pattern in concl_patterns:
        concl_match = re.search(pattern, text, re.DOTALL)
        if concl_match:
            content = concl_match.group(1).strip()
            if len(content) > 100 and is_conclusion_content(content):
                sections['conclusion'] = content
                break
    
    # More conservative fallback for literature review
    if 'literature review' not in sections:
        # Look for content between introduction and methodology that has literature review characteristics
        intro_pos = text.lower().find('introduction')
        method_pos = text.lower().find('methodology')
        if intro_pos != -1 and method_pos != -1 and method_pos > intro_pos:
            middle_section = text[intro_pos:method_pos]
            # Look for substantial content that looks like literature review
            paragraphs = middle_section.split('\n\n')
            lit_content = []
            for para in paragraphs:
                if len(para.strip()) > 200 and is_literature_review_content(para):
                    lit_content.append(para.strip())
            if lit_content:
                sections['literature review'] = '\n\n'.join(lit_content[:2])  # Take first 2 good paragraphs only
    
    # More conservative fallback for conclusion
    if 'conclusion' not in sections:
        # Look for content near the end that has conclusion characteristics
        doc_length = len(text)
        search_start = int(doc_length * 0.85)  # Start from 85% through the document
        end_section = text[search_start:]
        
        # Look for paragraphs that look like conclusion content
        paragraphs = end_section.split('\n\n')
        concl_content = []
        for para in paragraphs:
            if len(para.strip()) > 150 and is_conclusion_content(para):
                concl_content.append(para.strip())
        if concl_content:
            sections['conclusion'] = '\n\n'.join(concl_content[:1])  # Take only the first good paragraph
    
    # Special handling for literature review sections with Roman numerals
    if 'literature review' not in sections:
        # Look specifically for "II. Literature Review", "III. Literature Review", etc.
        roman_lit_patterns = [
            r'(?i)(?:II|III|IV|V)\.?\s+(?:literature review|related work|literature|review)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach|procedure))',
            r'(?i)(?:2|3|4|5)\.?\s+(?:literature review|related work|literature|review)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|experimental|approach|procedure))'
        ]
        
        for pattern in roman_lit_patterns:
            lit_match = re.search(pattern, text, re.DOTALL)
            if lit_match:
                content = lit_match.group(1).strip()
                if len(content) > 100:  # More lenient for Roman numeral sections
                    sections['literature review'] = content
                    print(f"Found literature review via Roman numeral pattern: {len(content)} characters")
                    break
    
    # Special handling for conclusion sections with Roman numerals
    if 'conclusion' not in sections:
        # Look specifically for "VI. Conclusion", "VII. Conclusion", etc.
        roman_concl_patterns = [
            r'(?i)(?:VI|VII|VIII|IX|X|XI|XII)\.?\s+(?:conclusion|conclusions)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
            r'(?i)(?:6|7|8|9|10|11|12)\.?\s+(?:conclusion|conclusions)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))'
        ]
        
        for pattern in roman_concl_patterns:
            concl_match = re.search(pattern, text, re.DOTALL)
            if concl_match:
                content = concl_match.group(1).strip()
                if len(content) > 100:  # More lenient for Roman numeral sections
                    sections['conclusion'] = content
                    print(f"Found conclusion via Roman numeral pattern: {len(content)} characters")
                    break
    
    return sections

def parse_pdf(pdf_bytes: bytes, filename: str) -> dict:
    text = extract_text_from_pdf_bytes(pdf_bytes)
    text = normalize_text(text)
    meta = find_title_and_authors(text)
    
    # Try heading-based extraction first
    sections = split_sections(text)
    
    # Debug: print what sections were found
    print(f"Found sections in {filename}: {list(sections.keys())}")
    
    # Always try content-pattern-based extraction as a supplement
    pattern_sections = extract_sections_by_content_patterns(text)
    # Merge pattern-based sections with heading-based sections (pattern-based takes precedence)
    for key, value in pattern_sections.items():
        if value and len(value) > 50:  # Only add if substantial content
            sections[key] = value
    
    # Debug: print final sections
    print(f"Final sections for {filename}: {list(sections.keys())}")

    # Extract sections with validation
    abstract = pick_section_content(sections, ["abstract", "summary"])
    introduction = pick_section_content(sections, ["introduction", "background", "overview"])
    
    # Validate methodology content to avoid random content
    methodology = pick_section_content(sections, ["methodology", "methods", "materials and methods", "experimental", "approach", "procedure", "method"])
    if methodology and not is_methodology_content(methodology):
        print(f"Methodology rejected (random content): {methodology[:200]}...")
        methodology = ""  # Clear if it doesn't look like methodology
    
    # Literature review section removed as requested
    literature_review = ""
    
    # Validate conclusion content - be more lenient for Roman numeral sections
    conclusion = pick_section_content(sections, ["conclusion", "conclusions", "discussion", "summary", "implications", "findings"])
    if conclusion:
        # Check if it's from a Roman numeral section - be more lenient
        is_roman_section = any(roman in conclusion for roman in ['VI.', 'VII.', 'VIII.', 'IX.', 'X.', 'XI.', 'XII.'])
        if is_roman_section:
            # For Roman numeral sections, just check minimum length
            if len(conclusion.strip()) < 50:
                print(f"Roman numeral conclusion too short: {conclusion[:100]}...")
                conclusion = ""
        else:
            # For regular sections, use strict validation
            if not is_conclusion_content(conclusion):
                print(f"Conclusion rejected: {conclusion[:200]}...")
                conclusion = ""
    
    references = pick_section_content(sections, ["references", "bibliography", "works cited"])

    data = {
        "file_name": filename,
        "title": meta.get("title"),
        "authors": meta.get("authors"),
        "abstract": abstract,
        "introduction": introduction,
        "methodology": methodology,
        "literature_review": literature_review,
        "conclusion": conclusion,
        "references": references,
    }
    return {k: clean_text_for_excel(v) for k, v in data.items()}

def create_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header = list(df.columns)
    ws.append(header)

    for row in df.itertuples(index=False, name=None):
        safe_row = [clean_text_for_excel(str(x) if x else "") for x in row]
        ws.append(safe_row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 40

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# ---------- Streamlit UI ----------
st.set_page_config(page_title="AI-Based Research Paper Extractor", layout="wide")

st.title("📘 CSIR  AI-Based Research Paper Extractor")
st.markdown("🔒 100% Local — No external API. Upload PDFs, extract structured research info, and download styled Excel.")

uploaded_files = st.file_uploader("📤 Upload one or more Research Papers (PDF)", type=["pdf"], accept_multiple_files=True)
if uploaded_files:
    if st.button("🚀 Click here to Extract and Generate Excel"):
        progress = st.progress(0)
        status = st.empty()
        results = []
        for i, file in enumerate(uploaded_files, start=1):
            try:
                status.info(f"Processing {file.name}...")
                result = parse_pdf(file.read(), file.name)
                results.append(result)
                status.success(f"✅ Extracted {file.name}")
            except Exception as e:
                st.error(f"❌ Failed to process {file.name}: {e}")
                st.text(traceback.format_exc())
            progress.progress(int(i / len(uploaded_files) * 100))
            time.sleep(0.2)

        if results:
            df = pd.DataFrame(results)
            st.success(f"🎉 Extraction complete for {len(results)} file(s)!")
            st.metric("📄 Total Papers", len(df))
            
            # Show detailed extraction statistics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📝 Abstracts Found", df['abstract'].notna().sum())
            with col2:
                st.metric("🔬 Methodologies Found", df['methodology'].notna().sum())
            with col3:
                st.metric("🎯 Conclusions Found", df['conclusion'].notna().sum())

            with st.expander("📊 Preview of Extracted Data"):
                st.dataframe(df, use_container_width=True)
            
            # Show section extraction details for debugging
            with st.expander("🔍 Section Extraction Details"):
                for i, result in enumerate(results):
                    st.write(f"**File {i+1}: {result['file_name']}**")
                    sections_found = []
                    for key in ['abstract', 'introduction', 'methodology', 'conclusion', 'references']:
                        if result.get(key) and len(str(result[key]).strip()) > 50:
                            sections_found.append(f"✅ {key.title()}")
                        else:
                            sections_found.append(f"❌ {key.title()}")
                    st.write(" | ".join(sections_found))
                    
                    # Show what sections were actually found in the document
                    st.write("**Sections found in document:**")
                    # This would require access to the sections dict, let's add a debug mode
                    if st.checkbox("Show debug info", key=f"debug_{i}"):
                        st.write("Debug information would be shown here...")
                    
                    # Show character counts for each section
                    st.write("**Character counts:**")
                    char_counts = []
                    for key in ['abstract', 'introduction', 'methodology', 'conclusion', 'references']:
                        count = len(str(result.get(key, '')).strip())
                        char_counts.append(f"{key.title()}: {count}")
                    st.write(" | ".join(char_counts))
                    
                    # Show validation status for methodology and conclusion
                    if result.get('methodology'):
                        method_content = str(result.get('methodology', ''))
                        is_valid_method = is_methodology_content(method_content)
                        st.write(f"**Methodology Validation:** {'✅ Valid' if is_valid_method else '❌ Invalid (random content)'}")
                        if not is_valid_method:
                            st.write(f"*Content preview: {method_content[:200]}...*")
                    
                    if result.get('conclusion'):
                        concl_content = str(result.get('conclusion', ''))
                        is_valid_concl = is_conclusion_content(concl_content)
                        st.write(f"**Conclusion Validation:** {'✅ Valid' if is_valid_concl else '❌ Invalid (random content)'}")
                        if not is_valid_concl:
                            st.write(f"*Content preview: {concl_content[:200]}...*")
                        else:
                            # Check if it's from a Roman numeral section
                            if any(roman in concl_content for roman in ['VI.', 'VII.', 'VIII.', 'IX.', 'X.', 'XI.', 'XII.']):
                                st.write("*✅ Found via Roman numeral heading (VI. Conclusion, etc.)*")
                    
                    st.write("---")

            excel_bytes = create_excel(df)
            st.download_button("⬇️ Download Styled Excel", excel_bytes, "research_papers_summary.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.balloons()
